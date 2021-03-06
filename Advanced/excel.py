from openpyxl import Workbook
wb = Workbook() # type: Workbook

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to celis
ws["A1"] = 42

#Rows can also be opended
ws.append([1 ,2, 3])

# Python types will automatically be converted
import datetime
ws["A2"] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")